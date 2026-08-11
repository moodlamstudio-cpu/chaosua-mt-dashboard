from openpyxl import load_workbook
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
ws=wb["Dashboard Calc"]
rows=list(ws.iter_rows(values_only=True))
for idx,r in enumerate(rows[:18]):
    s=[str(x) for x in r]
    if 'Month' in s or (len(r)>1 and r[0] in (1,2,3,4)):
        print(f"row[{idx}]:", s[:12])
wb.close()
