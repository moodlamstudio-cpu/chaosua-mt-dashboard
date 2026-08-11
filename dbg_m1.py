from openpyxl import load_workbook
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
ws=wb["Dashboard Calc"]
mon={}
raw_m1=None
for r in ws.iter_rows(values_only=True):
    v=r[0]
    if isinstance(v,(int,float)) and 1<=v<=12:
        mon[int(v)]={"act":(r[1] or 0)+(r[3] or 0),"le":(r[2] or 0)+(r[4] or 0),"ly":(r[5] or 0)+(r[6] or 0),"aop":(r[7] or 0)+(r[8] or 0)}
        if int(v)==1: raw_m1=(r[1],r[3],r[5],r[6],r[7],r[8])
print("M1 dict:",mon.get(1))
print("M1 raw:",raw_m1)
print("len mon:",len(mon))
# sum คะ
act=sum(mon[m]["act"] for m in range(1,9))
print("ACT sum raw baht:",act)
print("ACT sum MB:",round(act/1e6,2))
wb.close()
