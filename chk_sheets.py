from openpyxl import load_workbook
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
print("=== Sheets ทั้งหมด ===")
for s in wb.sheetnames: print(" -",s)
# ดู Dashboard Calc columns header (แถว 4)
ws=wb["Dashboard Calc"]
rows=list(ws.iter_rows(values_only=True))
print("\n=== Dashboard Calc แถวที่ 3-4 (header) ===")
for r in rows[2:4]:
    print([str(x) for x in r[:12]])
wb.close()
