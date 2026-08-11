from openpyxl import load_workbook
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
ws=wb["Dashboard Calc"]
rows=list(ws.iter_rows(values_only=True))
hdr=rows[3]
print("=== YTD row (หา 'YTD') ===")
for r in rows:
    if r[0]=="YTD":
        for ci in range(len(hdr)):
            if hdr[ci] and "None" not in str(r[ci]):
                print(f"  {hdr[ci]:16} = {r[ci]}")
# แสดง every month 12 รายการ col ที่มีค่าไม่ใช่ None
print("\n=== ทุกเดือน (col: ActualTotal, LYTotal, LETotal, AOP Total) ===")
for r in rows:
    if isinstance(r[0],(int,float)) and r[0] in range(1,13):
        a=r[9]; ly=r[10]; le=r[11]
        # หา AOP Total col
        aop=None
        for ci in range(len(r)):
            if hdr[ci] and 'AOP' in str(hdr[ci]) and ci<len(r):
                aopv=r[ci]
                if aopv and 'Total' in str(hdr[ci]): aop=aopv
        print(f"  M{r[0]}: Actual={a} LY={ly} LE={le}")
# ตรวจ LE column เพิ่มเติม: Makro Actual vs Makro LE ต่างกันไหมทั้ง 12
print("\n=== เปรียบเทียบ Makro Actual vs LE, Lotus Actual vs LE ===")
for r in rows:
    if isinstance(r[0],(int,float)) and r[0] in range(1,13):
        same = (r[1] and r[2] and abs(r[1]-r[2])<1e-6)
        sameL = (r[3] and r[4] and abs(r[3]-r[4])<1e-6)
        print(f"  M{r[0]}: MakroAct={r[1]} MakroLE={r[2]} same={same} | LotusAct={r[3]} LotusLE={r[4]} same={sameL}")
wb.close()
