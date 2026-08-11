from openpyxl import load_workbook
import json, datetime
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,data_only=True)
ws=wb["Dashboard Calc"]
rows=list(ws.iter_rows(min_row=4, max_row=17, values_only=True))
hdr=[(str(c).strip() if c is not None else "") for c in rows[0]]
def key(r,k): return r[hdr.index(k)] if k in hdr else None
months=[]; ytd=None
for r in rows[1:]:
    mn=r[0]
    if str(mn).strip().upper()=="YTD":
        ytd=r; continue
    months.append({
      "m":int(float(mn)),"label":{1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}[int(float(mn))],
      "makro":{"le":key(r,"Makro LE"),"ly":key(r,"Makro LY"),"aop":key(r,"Makro AOP")},
      "lotus":{"le":key(r,"Lotus LE"),"ly":key(r,"Lotus LY"),"aop":key(r,"Lotus AOP")},
      "total":{"le":key(r,"LE Total"),"ly":key(r,"LY Total"),"aop":key(r,"AOP Total")}
    })
def d(v): return round(float(v),1) if v not in (None,"") and not (isinstance(v,str) and '#' in v) else None
out={
 "title":"MT Sales Dashboard",
 "subtitle":"Lotus's + Makro (MT) · ยอดขายรายเดือน",
 "year":2026,"current_month":8,"unit":"ล้านบาท",
 "months":months,
 "ytd":{"makro_ly":d(key(ytd,"Makro LY")),"makro_aop":d(key(ytd,"Makro AOP")),"makro_le":d(key(ytd,"Makro LE")),
         "lotus_ly":d(key(ytd,"Lotus LY")),"lotus_aop":d(key(ytd,"Lotus AOP")),"lotus_le":d(key(ytd,"Lotus LE")),
         "total_le":d(key(ytd,"LE Total")),"total_ly":d(key(ytd,"LY Total")),"total_aop":d(key(ytd,"AOP Total"))},
 "updated":datetime.date.today().strftime("%Y-%m-%d")
}
path="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Web/LotusMakro-Dashboard/data.json"
with open(path,"w",encoding="utf-8") as fp: json.dump(out,fp,ensure_ascii=False,indent=2)
print("SAVED data.json",len(open(path,encoding="utf-8").read()),"bytes")
print("months:",len(months),"| YTD total LE:",out["ytd"]["total_le"],"LY:",out["ytd"]["total_ly"],"AOP:",out["ytd"]["total_aop"])
print("Makro LE YTD:",out["ytd"]["makro_le"],"Lotus LE YTD:",out["ytd"]["lotus_le"])
wb.close()
