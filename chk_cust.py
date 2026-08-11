from openpyxl import load_workbook
from collections import defaultdict
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
# 1) ดูชื่อ Customer ที่มีจริงใน Raw data
cust=defaultdict(float)
tot=0.0
for i,r in enumerate(wb["Raw data"].iter_rows(values_only=True)):
    if i==0: continue
    if r[0]!=2026: continue
    m=r[1]
    if m is None or m>8: continue
    c=r[4]
    val=r[9] or 0
    key=str(c)
    cust[key]+=val
    tot+=val
print("=== Raw data 2026 M1-8: แยก Customer ทั้งหมด ===")
for c,v in sorted(cust.items(),key=lambda x:-x[1]):
    print(f"  {c!r:30} {v/1e6:9.2f} MB")
print(f"  TOTAL raw M1-8 = {tot/1e6:.2f} MB")
# 2) ดู Customer ชื่อเดียวที่รวมได้ 104.3
print("\nช่องทางที่เข้าข่าย MT:",[c for c in cust if 'makro' in str(c).lower() or 'lotus' in str(c).lower()])
wb.close()
