from openpyxl import load_workbook
import json, datetime
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,data_only=True)
ws=wb["Dashboard Calc"]
rows=list(ws.iter_rows(min_row=4, max_row=17, values_only=True))
hdr=[(str(c).strip() if c is not None else "") for c in rows[0]]
print("COLUMNS:")
for j,h in enumerate(hdr): print(f"  [{j}] {h}")
# แสดงค่า month 1,8 และ YTD สำหรับทุกคอลัมน์หลัก
keys=["Month","Makro Actual","Makro LE","Lotus Actual","Lotus LE","Makro LY","Lotus LY","Makro AOP","Lotus AOP","Actual Total","LY Total","LE Total","AOP Total"]
for m in [0,7,12]:  # month1, month8, YTD index
    r=rows[m+1]
    print(f"\n-- row {rows[m+1][0]} --")
    for j,k in enumerate(keys):
        if k in hdr:
            val=r[hdr.index(k)]
            print(f"   {k}: {val}")
wb.close()
