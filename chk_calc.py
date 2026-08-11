from openpyxl import load_workbook
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
ws=wb["Dashboard Calc"]
print("=== Dashboard Calc YTD (แถว YTD) ===")
for i,r in enumerate(ws.iter_rows(values_only=True)):
    if r[0]=="YTD":
        print(f"  Makro Actual={r[1]} (ร.1)")
        print(f"  Lotus  Actual={r[3]} (ร.3)")
        print(f"  Actual Total={r[9]} (ร.9)")
# ตรวจ Raw data: หาว่า column ไหนคือ value - ลอง col ต่าง
# ดึงแถวแรก non-header ของ Raw data
ws2=wb["Raw data"]
hdr=None
for i,r in enumerate(ws2.iter_rows(values_only=True)):
    if i==0: hdr=r; break
print("\nRaw header:",[str(x) for x in hdr])
# หาผลรวม Makro ใน Raw by test ซึ่ง column คือ Net value
# ลองหา pattern: sum คอลัมน์ทั้งหมดที่มีตัวเลข per row ดู
ws3=wb["Raw data"]
samples=list(ws3.iter_rows(values_only=True))
# ดูแถวข้อมูลจริง 1 แถว (ที่ Makro)
for r in samples[1:5]:
    print("\nrow:",[str(x)[:12] for x in r])
wb.close()
