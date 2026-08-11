from openpyxl import load_workbook
import json, datetime
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,data_only=True)
ws=wb["Dashboard Calc"]
rows=list(ws.iter_rows(min_row=4, max_row=17, values_only=True))
hdr=[ (str(c).strip() if c is not None else "") for c in rows[0] ]
def conv(v):
    if isinstance(v,(datetime.datetime,datetime.date)):
        return str(v.date() if hasattr(v,"date") else v)
    if v is None: return None
    if isinstance(v,str) and ("#N/A" in v or "#NA" in v): return None
    if isinstance(v,(int,float)): return round(float(v),2)
    return v
data=[]
for r in rows[1:]:
    d={}
    for j in range(min(len(r),len(hdr))):
        d[hdr[j] or f"c{j}"]=conv(r[j])
    data.append(d)
out={"year":2026,"latest_month":8,"headers":hdr,"months":data}
path="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Web/LotusMakro-Dashboard/data.json"
with open(path,"w",encoding="utf-8") as fp:
    json.dump(out,fp,ensure_ascii=False,indent=2)
print("SAVED:",path,"bytes:",len(open(path,encoding='utf-8').read()))
# ยอด YTD แสดงตัวอย่าง
ytd=data[-1]
print("YTD sample:", {k:ytd[k] for k in ["Makro Actual","Lotus Actual","Actual Total","LY Total","AOP Total","Makro LY","Lotus LY","Makro AOP","Lotus AOP"]})
wb.close()
