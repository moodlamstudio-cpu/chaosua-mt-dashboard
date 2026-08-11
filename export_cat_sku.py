from openpyxl import load_workbook
from collections import defaultdict
import json, datetime
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
# Product Master map
pm={}
ws=wb["Product Master"]
for i,r in enumerate(ws.iter_rows(values_only=True)):
    if i==0: continue
    mat=str(r[0])
    pm[mat]={"d":(r[4] or ""),"g":(r[8] or ""),"c1":(r[13] or "")}
# Raw: YTD 2026 (ม.1-8) เฉพาะ Makro/Lotus
ws2=wb["Raw data"]
cat_val=defaultdict(float); sku_val=defaultdict(float)
for i,r in enumerate(ws2.iter_rows(values_only=True)):
    if i==0: continue
    if r[0]!=2026 or r[1] is None or r[1]>8: continue
    mm=str(r[4] or "").replace(" ","").lower()
    if mm not in ("makro","lotus's","lotus"): continue
    mat=str(r[5]); val=r[9] or 0
    info=pm.get(mat,{})
    cat=info.get("c1") or info.get("g") or "Other"
    cat_val[cat]+=val
    if info.get("d"): sku_val[info["d"]]+=val
# เขียนลง data.json
dp="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Web/LotusMakro-Dashboard/data.json"
d=json.load(open(dp,encoding="utf-8"))
# category: เรียง Top 12 + Other
cats=[{"name":k,"mb":round(v/1e6,2)} for k,v in sorted(cat_val.items(),key=lambda x:-x[1]) if v>0]
tot=sum(v/1e6 for _,v in cat_val.items())
if len(cats)>12:
    top=cats[:11]; rest=round(sum(c["mb"] for c in cats[11:]),2)
    cats=top+[{"name":"Other","mb":rest}]
tot=round(sum(c["mb"] for c in cats),2)
for c in cats: c["pct"]=round(c["mb"]/tot*100,1) if tot else 0
d["category"]={"total_mb":tot,"items":cats}
d["top_sku"]=[{"name":k,"mb":round(v/1e6,2)} for k,v in sorted(sku_val.items(),key=lambda x:-x[1])[:10]]
d["updated"]=datetime.date.today().strftime("%Y-%m-%d")
json.dump(d,open(dp,"w",encoding="utf-8"),ensure_ascii=False,indent=2)
print("WAIT done. category total MB:",tot)
for c in cats: print(f"  {c['mb']:7.2f} MB  {c['pct']:5.1f}%  {c['name']}")
print("Top SKU:")
for s in d["top_sku"]: print(f"  {s['mb']:7.2f} MB  {s['name']}")
wb.close()
