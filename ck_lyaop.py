from openpyxl import load_workbook
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
# 1) Dashboard Calc LY/AOP monthly -> หา YTD 12 เดือน
ws=wb["Dashboard Calc"]
rows=list(ws.iter_rows(values_only=True))
mon={}
for r in rows:
    if isinstance(r[0],(int,float)) and 1<=r[0]<=12:
        mon[r[0]]={"act":(r[1] or 0)+(r[3] or 0),"le":(r[2] or 0)+(r[4] or 0),"ly":(r[5] or 0)+(r[6] or 0),"aop":(r[7] or 0)+(r[8] or 0)}
ly12=sum(mon[m]["ly"] for m in range(1,13))
aop12=sum(mon[m]["aop"] for m in range(1,13))
ly8=sum(mon[m]["ly"] for m in range(1,9))
aop8=sum(mon[m]["aop"] for m in range(1,9))
le12=sum(mon[m]["le"] for m in range(1,13))
print(f"Dashboard Calc: LY12={ly12:.2f} AOP12={aop12:.2f} LY8={ly8:.2f} AOP8={aop8:.2f} LE12={le12:.2f}")
# monthly LY & AOP ทั้ง 12
print("LY monthly:",[round(mon[m]['ly'],2) for m in range(1,13)])
print("AOP monthly:",[round(mon[m]['aop'],2) for m in range(1,13)])
# 2) Data AOP Sale sheet
ws2=wb["Data AOP Sale"]
print("\n=== Data AOP Sale (first 6 rows) ===")
for i,r in enumerate(ws2.iter_rows(values_only=True)):
    if i>5: break
    print([str(x)[:12] for x in r])
wb.close()
