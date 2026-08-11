from openpyxl import load_workbook
from collections import defaultdict
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
ws=wb["Raw data"]
# sum อย่างเดียว: Makro + Tesco, year=2026, month<=8
tot=0.0; mak=0.0; tes=0.0
for i,r in enumerate(ws.iter_rows(values_only=True)):
    if i==0: continue
    if r[0]!=2026: continue
    m=r[1]
    if m is None or m>8: continue
    c=str(r[4]).lower()
    v=r[9] or 0
    if c=="makro": mak+=v; tot+=v
    elif c=="tesco": tes+=v; tot+=v
print(f"Makro = {mak/1e6:.2f}")
print(f"Tesco(Lotus) = {tes/1e6:.2f}")
print(f"รวม Makro+Tesco = {tot/1e6:.2f}  (Dashboard Calc = 104.35)")
wb.close()
