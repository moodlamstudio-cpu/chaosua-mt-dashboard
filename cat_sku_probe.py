from openpyxl import load_workbook
from collections import defaultdict
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
# อ่าน Product Master: material -> desc, cat, group_desc
pm={}
ws=wb["Product Master"]
hdr=None
for i,r in enumerate(ws.iter_rows(values_only=True)):
    if i==0: hdr=[str(x).strip() if x else "" for x in r]; continue
    mat=r[0]; desc=r[4]; grp=r[7]; grpd=r[8]; cat0=r[12]; cat1=r[13]; cat2=r[14]; cat3=r[15]
    pm[str(mat)]={"d":desc,"g":grpd,"c1":cat1,"c0":cat0,"mgrp":grp}
# ดูตัวอย่าง category ต่างกันคร่าว
cats=defaultdict(int)
for k,v in pm.items():
    cats[str(v["c1"] or v["g"] or v["c0"])]+=1
print("=== Category ตัวอย่าง (Count) ===")
for c,n in sorted(cats.items(),key=lambda x:-x[1])[:20]:
    print(f"  {n:5}  {c}")
# Raw: ยอด 2026 YTD แยก category (เฉพาะ Makro/Lotus) + SKU top
ws2=wb["Raw data"]
cat_val=defaultdict(float); sku_val=defaultdict(lambda:[0.0,None]) # val, desc
mt_target={"makro","lotus`s","lotus","makro"}
cust_lookup={c.lower().replace(" ",""):c for c in ["Makro","Lotus's"]}
total=0
for i,r in enumerate(ws2.iter_rows(values_only=True)):
    if i==0: continue
    yr=r[0]; mo=r[1]; cust=str(r[4] or "").lower()
    mm=cust.replace(" ","")
    if yr!=2026 or mo is None or mo>8: continue
    mat=str(r[5]); val=r[9] or 0
    if mm not in ("makro","lotus's","lotus"): 
        continue
    info=pm.get(mat,{})
    cat=info.get("c1") or info.get("g") or (info.get("d") or "Other")
    cat_val[cat]+=val
    if info.get("d"):
        sku_val[mat][0]+=val
        if sku_val[mat][1] is None: sku_val[mat][1]=info["d"]
    total+=val
print(f"\n=== YTD2026 (ม.1-8) Makro+Lotus total = {total/1e6:.1f} MB ===")
print("=== Category (MB) ===")
for c,v in sorted(cat_val.items(),key=lambda x:-x[1])[:15]:
    print(f"  {v/1e6:8.2f} MB  {c}")
print("=== Top 10 SKU (MB) ===")
for mat,(v,d) in sorted(sku_val.items(),key=lambda x:-x[1][0])[:10]:
    print(f"  {v/1e6:8.2f} MB  {d}")
wb.close()
