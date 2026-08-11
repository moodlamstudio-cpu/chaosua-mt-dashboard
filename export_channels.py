from openpyxl import load_workbook
from collections import defaultdict
import json
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
# Product Master
pm={}
for i,r in enumerate(wb["Product Master"].iter_rows(values_only=True)):
    if i==0 or r[0] is None: continue
    pm[str(r[0])]={"d":(r[4] or ""),"c1":(r[13] or ""),"g":(r[8] or "")}

# Channel map: Tesco->Lotus' (Raw ใช้ชื่อ Tesco สำหรับ Lotus)
def norm(c):
    s=str(c).strip().lower()
    if s=="makro": return "MAKRO"
    if s=="tesco": return "LOTUS'"   # ใน Raw data Tesco = Lotus's
    return str(c).strip()

# Accumulate per (channel, month, material)
# channel->month->val  และ channel->material->[val,desc,cat]
ch_m = defaultdict(lambda: defaultdict(float))
ch_mat = defaultdict(lambda: dict())
ch_cat_series = defaultdict(lambda: defaultdict(lambda: {"actual":defaultdict(float),"ly":defaultdict(float)}))
ch_sku_series = defaultdict(lambda: defaultdict(lambda: {"actual":defaultdict(float),"ly":defaultdict(float),"cat":""}))
ch_annual = defaultdict(lambda: defaultdict(float))
ch_annual_cat = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
ch_annual_sku = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
ch_history = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
ch_history_cat = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(float))))
ch_history_sku = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(float))))
Y=2026; MMAX=8
n_rows=0
for i,r in enumerate(wb["Raw data"].iter_rows(values_only=True)):
    if i==0: continue
    if not isinstance(r[0],(int,float)): continue
    m=r[1]
    if m is None or not 1<=m<=12: continue
    ch=norm(r[4]); mat=str(r[5]); val=r[9] or 0
    info=pm.get(mat,{})
    cat=info.get("c1") or info.get("g") or "Other"
    d=info.get("d") or mat
    year=int(r[0])
    ch_annual[ch][year]+=val
    ch_annual_cat[ch][cat][year]+=val
    ch_annual_sku[ch][d][year]+=val
    ch_history[ch][year][m]+=val
    ch_history_cat[ch][cat][year][m]+=val
    ch_history_sku[ch][d][year][m]+=val
    if year not in (Y,Y-1): continue
    if year==Y and m>MMAX: continue
    series_key="actual" if year==Y else "ly"
    ch_cat_series[ch][cat][series_key][m]+=val
    ch_sku_series[ch][d][series_key][m]+=val
    ch_sku_series[ch][d]["cat"]=cat
    if r[0]!=Y: continue
    ch_m[ch][m]+=val
    e=ch_mat[ch].get(mat)
    if e is None: ch_mat[ch][mat]=[val,d,cat]
    else: e[0]+=val
n_total_rows=sum(len(v) for v in ch_mat.values())
# Build per channel JSON
ALL=[]
chan_data={}
for ch in ch_m:
    mats=sorted(ch_mat[ch].items(),key=lambda x:-x[1][0])
    tot=sum(v[0] for v in [x[1] for x in mats])
    # category
    catt=defaultdict(float)
    for mat,(v,d,c) in ch_mat[ch].items(): catt[c]+=v
    catitems=[{"name":c,"mb":round(v/1e6,2),"pct":round(v/tot*100,1) if tot else 0} for c,v in sorted(catt.items(),key=lambda x:-x[1])]
    sku_all=[{"name":d,"mb":round(v/1e6,2),"cat":c} for mat,(v,d,c) in mats]
    sku_bycat={c:[s for s in sku_all if s["cat"]==c][:10] for c,_ in catt.items()}
    chan_data[ch]={"label":ch,"total_mb":round(tot/1e6,2),"monthly":{str(mm):round(ch_m[ch][mm]/1e6,2) for mm in range(1,13)},
      "category":{"total_mb":round(tot/1e6,2),"items":catitems},"top_sku_all":sku_all[:15],"top_sku_by_cat":sku_bycat,
      "monthly_by_cat":{cat:{k:{str(mm):round(v[k][mm]/1e6,2) for mm in range(1,13)} for k in ("actual","ly")} for cat,v in ch_cat_series[ch].items()},
      "monthly_by_sku":{name:{"cat":v["cat"],**{k:{str(mm):round(v[k][mm]/1e6,2) for mm in range(1,13)} for k in ("actual","ly")}} for name,v in ch_sku_series[ch].items()},
      "annual":{str(y):round(v/1e6,2) for y,v in sorted(ch_annual[ch].items())},
      "annual_by_cat":{cat:{str(y):round(v/1e6,2) for y,v in sorted(years.items())} for cat,years in ch_annual_cat[ch].items()},
      "annual_by_sku":{name:{str(y):round(v/1e6,2) for y,v in sorted(years.items())} for name,years in ch_annual_sku[ch].items()},
      "history":{str(y):{str(mm):round(v/1e6,2) for mm,v in sorted(months.items())} for y,months in sorted(ch_history[ch].items())},
      "history_by_cat":{cat:{str(y):{str(mm):round(v/1e6,2) for mm,v in sorted(months.items())} for y,months in sorted(years.items())} for cat,years in ch_history_cat[ch].items()},
      "history_by_sku":{name:{str(y):{str(mm):round(v/1e6,2) for mm,v in sorted(months.items())} for y,months in sorted(years.items())} for name,years in ch_history_sku[ch].items()}}
    ALL.append((tot,ch))
# MT = MAKRO + LOTUS'
mt=["MAKRO","LOTUS'"]
tot=sum(chan_data[c]["total_mb"]*1e6 for c in mt)
# merge monthly
m=defaultdict(float)
for c in mt:
    for mm,v in chan_data[c]["monthly"].items(): m[mm]+=v
# merge category
catm=defaultdict(float)
for c in mt:
    for it in chan_data[c]["category"]["items"]: catm[it["name"]]+=it["mb"]
catitems=[{"name":c,"mb":round(v,2),"pct":round(v*1e6/tot*100,1) if tot else 0} for c,v in sorted(catm.items(),key=lambda x:-x[1])]
# merge sku
skum=defaultdict(lambda:[0.0,"",''])
for c in mt:
    for mat,(v,d,cat) in ch_mat[c].items():
        skum[mat][0]+=v
        if skum[mat][1]=="": skum[mat][1]=d; skum[mat][2]=cat
sku_all=[{"name":d,"mb":round(v/1e6,2),"cat":c} for mat,(v,d,c) in skum.items()]
sku_all.sort(key=lambda x:-x["mb"])
chan_data["MT"]={"label":"MT (Makro+Lotus')","total_mb":round(tot/1e6,2),"monthly":{str(mm):round(m[mm]/1e6,2) for mm in range(1,13)},
  "category":{"total_mb":round(tot/1e6,2),"items":catitems},"top_sku_all":sku_all[:15],
  "top_sku_by_cat":{c:[s for s in sku_all if s["cat"]==c][:10] for c,_ in catm.items()}}
# merge monthly category/SKU series for MT
for field in ("monthly_by_cat","monthly_by_sku"):
    merged={}
    for channel_id in mt:
        for name,series in chan_data[channel_id][field].items():
            out=merged.setdefault(name,{"actual":{str(mm):0 for mm in range(1,13)},"ly":{str(mm):0 for mm in range(1,13)}})
            if "cat" in series: out["cat"]=series["cat"]
            for key in ("actual","ly"):
                for mm in range(1,13): out[key][str(mm)]=round(out[key][str(mm)]+series[key][str(mm)],2)
    chan_data["MT"][field]=merged
for field in ("annual","annual_by_cat","annual_by_sku"):
    merged={}
    for channel_id in mt:
        source=chan_data[channel_id][field]
        if field=="annual": source={"_total":source}
        for name,years in source.items():
            out=merged.setdefault(name,{})
            for year,value in years.items(): out[year]=round(out.get(year,0)+value,2)
    chan_data["MT"][field]=merged.get("_total",{}) if field=="annual" else merged
for field in ("history","history_by_cat","history_by_sku"):
    merged={}
    for channel_id in mt:
        source=chan_data[channel_id][field]
        if field=="history": source={"_total":source}
        for name,years in source.items():
            out=merged.setdefault(name,{})
            for year,months in years.items():
                yy=out.setdefault(year,{})
                for mm,value in months.items(): yy[mm]=round(yy.get(mm,0)+value,2)
    chan_data["MT"][field]=merged.get("_total",{}) if field=="history" else merged
# ใส่ ALL channels list (เรียงยอด) สำหรับ dropdown
allch=[c for _,c in sorted(ALL,reverse=True)]
chan_data["_channels"]=[{"id":"MT","label":"MT (Makro+Lotus')"}]+[{"id":c,"label":c} for c in allch]
json.dump(chan_data,open("data_channels.json","w",encoding="utf-8"),ensure_ascii=False,indent=2)
print("channels:",[c for _,c in sorted(ALL,reverse=True)])
print("MT total:",round(tot/1e6,2),"MB")
print("MT category:",[(i['name'],i['mb'],i['pct']) for i in catitems])
wb.close()
