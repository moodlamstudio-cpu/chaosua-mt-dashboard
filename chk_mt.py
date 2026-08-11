from openpyxl import load_workbook
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
# Dashboard MT - ดูโครงสร้าง
ws=wb["Dashboard MT"]
print("=== Dashboard MT (first 8 rows) ===")
for i,r in enumerate(ws.iter_rows(values_only=True)):
    if i>7: break
    print([str(x)[:14] for x in r[:10]])
# Pivot
ws2=wb["Pivot"]
print("\n=== Pivot (first 6 rows) ===")
for i,r in enumerate(ws2.iter_rows(values_only=True)):
    if i>5: break
    print([str(x)[:14] for x in r[:8]])
wb.close()
