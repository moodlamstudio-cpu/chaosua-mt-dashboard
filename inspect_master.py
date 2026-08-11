from openpyxl import load_workbook
from collections import defaultdict
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
# 1) Product Master - ดูว่ามี category column ไหม
print("=== Product Master ===")
ws=wb["Product Master"]
rows=ws.iter_rows(min_row=1,max_row=8,values_only=True)
hdr=next(rows,None); print("hdr:",hdr)
for r in rows: print("  ",r)
# 2) Raw data - ว่า Customer ไหนคือ Makro/Lotus + มี category mapping มั้ย ดูกลุ่ม Material
print("\n=== Raw data: เจอเฉพาะ Makro/Lotus ไหม + ตัวอย่าง material ===")
ws2=wb["Raw data"]
hdr2=None
mats=defaultdict(list)
cust_set=set()
for i,r in enumerate(ws2.iter_rows(values_only=True)):
    if i==0: hdr2=r; continue
    c=r[4] if r[4] else "?"
    cust_set.add(str(c))
    if len(mats)<=3:
        mats[str(r[5])].append((r[0],r[1],r[9]))
print("Customers ใน raw:",sorted(cust_set))
wb.close()
