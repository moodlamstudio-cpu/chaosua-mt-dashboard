from openpyxl import load_workbook
import json
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
ws=wb["Dashboard Calc"]
# header แถวที่ 4: col1=Makro_Actual,2=Makro_LE,3=Lotus_Actual,4=Lotus_LE,5=Makro_LY,6=Lotus_LY,7=Makro_AOP,8=Lotus_AOP,9=ActualTotal,10=LYTotal,11=LETotal,12=AOPTotal
rows=list(ws.iter_rows(values_only=True))
# หาแถว http: col0=Month
by_m={}
for r in rows:
    if r[0] in (1,2,3,4,5,6,7,8,9,10,11,12):
        by_m[r[0]]=r
# MT monthly: makro+lots for each metric
def mt_series(idx_makro, idx_lotus):
    out={m:0.0 for m in range(1,13)}
    for m,r in by_m.items():
        a=r[idx_makro] or 0; b=r[idx_lotus] or 0
        out[m]=a+b
    return out
LY=mt_series(5,6)   # Makro LE? ลำดับ: idx5=Makro_LY,6=Lotus_LY
LE=mt_series(2,4)   # Makro_LE idx2, Lotus_LE idx4
AOP=mt_series(7,8)  # Makro_AOP idx7, Lotus_AOP idx8
AC=mt_series(1,3)   # Makro_Actual idx1, Lotus_Actual idx3
print("MT LY monthly:",{m:round(LY[m]/1e6,2) for m in range(1,9)})
print("MT LE monthly:",{m:round(LE[m]/1e6,2) for m in range(1,13)})
print("MT AOP monthly:",{m:round(AOP[m]/1e6,2) for m in range(1,13)})
# YTD sum
ly=sum(LY.values())/1e6; le=sum(LE.values())/1e6; aop=sum(AOP.values())/1e6; ac=sum(AC.values())/1e6
print(f"YTD: Actual={ac:.2f} LE={le:.2f} LY={ly:.2f} AOP={aop:.2f}")
# เขียนลง data_channels.json -> MT
d=json.load(open("data_channels.json",encoding="utf-8"))
mt=d["MT"]
mt["actual_ytd"]=round(sum(AC.values())/1e6,2)
mt["le_ytd"]=round(sum(LE.values())/1e6,2) if False else round(le,2)
mt["leY"]={m:round(LE[m]/1e6,2) for m in range(1,13)}
mt["ly"]={m:round(LY[m]/1e6,2) for m in range(1,13)}
mt["aop"]={m:round(AOP[m]/1e6,2) for m in range(1,13)}
mt["ly_ytd"]=round(ly,2); mt["le_ytd"]=round(le,2); mt["aop_ytd"]=round(aop,2)
json.dump(d,open("data_channels.json","w",encoding="utf-8"),ensure_ascii=False,indent=2)
print("\nMT KPI: Actual",mt["actual_ytd"],"| LE",mt["le_ytd"],"| LY",mt["ly_ytd"],"| AOP",mt["aop_ytd"])
wb.close()
