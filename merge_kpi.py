from openpyxl import load_workbook
import json
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
ws=wb["Dashboard Calc"]
rows=list(ws.iter_rows(values_only=True))
# col: 1=MakroAct 2=MakroLE 3=LotusAct 4=LotusLE 5=MakroLY 6=LotusLY 7=MakroAOP 8=LotusAOP
mon={}
for r in rows:
    if isinstance(r[0],(int,float)) and r[0] in range(1,13):
        mon[r[0]]={"act":(r[1] or 0)+(r[3] or 0),"le":(r[2] or 0)+(r[4] or 0),"ly":(r[5] or 0)+(r[6] or 0),"aop":(r[7] or 0)+(r[8] or 0)}
# KPI: Actual M1-8 (actual), LE = Actual M1-7 + LE M8, LY = LY M1-8
act=sum(mon[m]["act"] for m in range(1,9))/1e6
le_mix=(sum(mon[m]["act"] for m in range(1,8))+mon[8]["le"])/1e6
ly=sum(mon[m]["ly"] for m in range(1,9))/1e6
aop=sum(mon[m]["aop"] for m in range(1,9))/1e6
print(f"MT KPI mix: Actual={act:.2f} LE(mix)={le_mix:.2f} LY={ly:.2f} AOP={aop:.2f} vsAOP={le_mix/aop*100:.1f}%")
# เขียนลง data.json
d=json.load(open("data_channels.json",encoding="utf-8"))
mt=d["MT"]
mt["actual_ytd"]=round(act,2)
mt["le_ytd_mix"]=round(le_mix,2)
mt["ly_ytd"]=round(ly,2)
mt["aop_ytd"]=round(aop,2)
# monthly series (MB) สำหรับ chart
mt["s_actual"]={str(m):round(mon[m]["act"]/1e6,2) for m in range(1,9)}
mt["s_le"]={str(m):round(mon[m]["le"]/1e6,2) for m in range(1,13)}
mt["s_ly"]={str(m):round(mon[m]["ly"]/1e6,2) for m in range(1,13)}
mt["s_aop"]={str(m):round(mon[m]["aop"]/1e6,2) for m in range(1,13)}
json.dump(d,open("data_channels.json","w",encoding="utf-8"),ensure_ascii=False,indent=2)
print("บันทึกแล้ว MT keys:",[k for k in mt if 's_' in k or k in ('actual_ytd','le_ytd_mix','ly_ytd','aop_ytd')])
wb.close()
