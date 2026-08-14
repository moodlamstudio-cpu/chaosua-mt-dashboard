# -*- coding: utf-8 -*-
"""
regenerate_webdata.py
Regenerate data_channels.json (the file the dashboard UI loads) from the source
workbook Sale Lotus Makro_Dashboard_Pivot.xlsx.

Faithfully reproduces the exact committed data_channels.json structure
(base channel layer + MT merge + LE/LY/AOP series + YTD KPIs) using the CURRENT
Raw data schema:

  Raw data columns:
    0 Year, 1 Month, 2 Week, 3 Date, 4 Channel Group (= "MT" for all rows),
    5 Customer (channel: Makro/Tesco/7Eleven/...), 6 Material, 7 Qty EA,
    8 Qty BOX, 9 Qty ton, 10 Net value THB   <-- NOTE value is col 10

Channel = Customer (col 5). "tesco" -> LOTUS', "makro" -> MAKRO.
Sales value = Net value THB (col 10).
LE/LY/AOP series + YTD KPIs come from the "Dashboard Calc" sheet (MB units).

This script only refreshes numbers; it does not change UI/business formulas.
"""
import json
import datetime
from collections import defaultdict
from openpyxl import load_workbook

SRC = r"C:\Users\teera\OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD\Desktop\Chaosua_Ice\Data\Sales report\Sale Lotus Makro_Dashboard_Pivot.xlsx"
OUT = "data_channels.json"

FISCAL_YEAR = 2026   # FISCAL year shown by the dashboard (unchanged)

def norm_channel(c):
    s = str(c).strip()
    low = s.lower()
    if low == "makro":
        return "MAKRO"
    if low in ("tesco", "lotus's", "lotus"):
        return "LOTUS'"
    return s

def init_channel():
    return {
        "annual": defaultdict(float),
        "annual_cat": defaultdict(lambda: defaultdict(float)),
        "annual_sku": defaultdict(lambda: defaultdict(float)),
        "history": defaultdict(lambda: defaultdict(float)),
        "history_cat": defaultdict(lambda: defaultdict(lambda: defaultdict(float))),
        "history_sku": defaultdict(lambda: defaultdict(lambda: defaultdict(float))),
        "monthly": defaultdict(float),
        "mat": defaultdict(lambda: [0.0, "", ""]),       # mat -> [val, desc, cat]
        "cat_series": defaultdict(lambda: defaultdict(lambda: defaultdict(float))),  # cat -> actual/ly -> month -> val
        "sku_series": defaultdict(lambda: defaultdict(lambda: defaultdict(float))),  # name-> actual/ly -> month -> val
        "sku_name": {},                                   # mat -> desc
        "sku_cat": {},                                    # mat -> cat
        "sku_cat_desc": {},                               # desc -> cat
    }

def load_product_master(wb):
    pm = {}
    ws = wb["Product Master"]
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or r[0] is None:
            continue
        pm[str(r[0])] = {"d": (r[4] or ""), "g": (r[8] or ""), "c1": (r[13] or "")}
    return pm

def build():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    pm = load_product_master(wb)
    ws = wb["Raw data"]

    ch_struct = defaultdict(init_channel)

    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        # Year (col 0), Month (col 1)
        try:
            year = int(r[0])
            month = int(r[1])
        except (TypeError, ValueError):
            continue
        if not (1 <= month <= 12):
            continue
        ch = norm_channel(r[5])          # Customer = channel
        mat = str(r[6])                  # Material
        try:
            val = float(r[10] or 0)      # Net value THB
        except (TypeError, ValueError):
            val = 0
        info = pm.get(mat, {}) or {}
        cat = info.get("c1") or info.get("g") or "Other"
        desc = info.get("d") or mat
        s = ch_struct[ch]

        s["annual"][year] += val
        s["annual_cat"][cat][year] += val
        s["annual_sku"][desc][year] += val
        s["history"][year][month] += val
        s["history_cat"][cat][year][month] += val
        s["history_sku"][desc][year][month] += val

        if year == FISCAL_YEAR:
            s["monthly"][month] += val
            if month <= 12:
                s["cat_series"][cat]["actual"][month] += val
                s["sku_series"][desc]["actual"][month] += val
                s["sku_name"][mat] = desc
                s["sku_cat"][mat] = cat
                s["sku_cat_desc"][desc] = cat
                e = s["mat"][mat]
                e[0] += val
                if not e[1]:
                    e[1] = desc
                    e[2] = cat
        elif year == FISCAL_YEAR - 1:
            if month <= 12:
                s["cat_series"][cat]["ly"][month] += val
                s["sku_series"][desc]["ly"][month] += val
    wb.close()

    chan_data = {}
    totals = []
    for ch, s in ch_struct.items():
        # Write every channel that has any sales data (current OR historical), so the
        # all-channel annual totals reconcile to the workbook. Channels are only added
        # to the UI picker (_channels) when they have current-year (2026) activity.
        total_all = sum(v for v in s["annual"].values())
        if total_all <= 0:
            continue
        mats = sorted(s["mat"].items(), key=lambda x: -x[1][0])
        tot = sum(v[0] for v in (x[1] for x in mats))
        # category (current year)
        catt = defaultdict(float)
        for mat, (v, d, c) in s["mat"].items():
            catt[c] += v
        catitems = [{"name": c, "mb": round(v / 1e6, 2),
                     "pct": round(v / tot * 100, 1) if tot else 0}
                    for c, v in sorted(catt.items(), key=lambda x: -x[1])]
        sku_all = [{"name": d, "mb": round(v / 1e6, 2), "cat": c}
                   for mat, (v, d, c) in mats]
        sku_bycat = {c: [x for x in sku_all if x["cat"] == c] for c in catt}
        monthly_by_cat = {}
        for cat, series in s["cat_series"].items():
            monthly_by_cat[cat] = {
                "actual": {str(mm): round(series["actual"][mm] / 1e6, 2) for mm in range(1, 13)},
                "ly": {str(mm): round(series["ly"][mm] / 1e6, 2) for mm in range(1, 13)},
            }
        # Resolve each SKU's category by description, with a material-key reverse
        # fallback (sku_cat is keyed by material). This fixes category-filtered
        # SKU tables showing empty categories.
        def sku_category(desc):
            c = s["sku_cat_desc"].get(desc)
            if c:
                return c
            for mat, d in s["sku_name"].items():
                if d == desc:
                    return s["sku_cat"].get(mat, "")
            return ""
        monthly_by_sku = {}
        for desc, series in s["sku_series"].items():
            monthly_by_sku[desc] = {
                "cat": sku_category(desc),
                "actual": {str(mm): round(series["actual"][mm] / 1e6, 2) for mm in range(1, 13)},
                "ly": {str(mm): round(series["ly"][mm] / 1e6, 2) for mm in range(1, 13)},
            }
        chan_data[ch] = {
            "label": ch,
            "total_mb": round(tot / 1e6, 2),
            "monthly": {str(mm): round(s["monthly"][mm] / 1e6, 2) for mm in range(1, 13)},
            "category": {"total_mb": round(tot / 1e6, 2), "items": catitems},
            "top_sku_all": sku_all,
            "top_sku_by_cat": sku_bycat,
            "monthly_by_cat": monthly_by_cat,
            "monthly_by_sku": monthly_by_sku,
            "annual": {str(y): round(v / 1e6, 2) for y, v in sorted(s["annual"].items())},
            "annual_by_cat": {cat: {str(y): round(v / 1e6, 2) for y, v in sorted(ys.items())}
                              for cat, ys in s["annual_cat"].items()},
            "annual_by_sku": {name: {str(y): round(v / 1e6, 2) for y, v in sorted(ys.items())}
                              for name, ys in s["annual_sku"].items()},
            "history": {str(y): {str(m): round(v / 1e6, 2) for m, v in sorted(ms.items())}
                        for y, ms in sorted(s["history"].items())},
            "history_by_cat": {cat: {str(y): {str(m): round(v / 1e6, 2) for m, v in sorted(ms.items())}
                                    for y, ms in sorted(ys.items())}
                               for cat, ys in s["history_cat"].items()},
            "history_by_sku": {name: {str(y): {str(m): round(v / 1e6, 2) for m, v in sorted(ms.items())}
                                     for y, ms in sorted(ys.items())}
                               for name, ys in s["history_sku"].items()},
        }
        totals.append((tot, ch))

    # ---- MT = MAKRO + LOTUS' (merges all structures) ----
    mt = ["MAKRO", "LOTUS'"]
    def add_num(a, b):
        return round((a or 0) + (b or 0), 2)

    mt_channel = {
        "label": "MT (Makro+Lotus')",
        "monthly": {str(m): add_num(*[chan_data[c]["monthly"][str(m)] for c in mt]) for m in range(1, 13)},
    }
    tot = sum(chan_data[c]["total_mb"] * 1e6 for c in mt)
    mt_channel["total_mb"] = round(tot / 1e6, 2)

    # category merge
    catm = defaultdict(float)
    for c in mt:
        for it in chan_data[c]["category"]["items"]:
            catm[it["name"]] += it["mb"]
    catitems = [{"name": cc, "mb": round(v, 2),
                 "pct": round(v * 1e6 / tot * 100, 1) if tot else 0}
                for cc, v in sorted(catm.items(), key=lambda x: -x[1])]
    mt_channel["category"] = {"total_mb": round(tot / 1e6, 2), "items": catitems}

    # sku merge
    skum = defaultdict(lambda: [0.0, "", ""])
    for c in mt:
        for e in chan_data[c]["top_sku_all"]:
            e0 = skum[e["name"]]
            e0[0] += e["mb"]
            if not e0[1]:
                e0[1] = e["name"]; e0[2] = e["cat"]
    sku_all = [{"name": v[1], "mb": round(v[0], 2), "cat": v[2]} for v in skum.values()]
    sku_all.sort(key=lambda x: -x["mb"])
    mt_channel["top_sku_all"] = sku_all
    mt_channel["top_sku_by_cat"] = {c: [s for s in sku_all if s["cat"] == c] for c in catm}

    # merge focus series (monthly_by_cat / monthly_by_sku)
    for field in ("monthly_by_cat", "monthly_by_sku"):
        merged = {}
        for ch_id in mt:
            for name, series in chan_data[ch_id][field].items():
                out = merged.setdefault(name, {"actual": {str(m): 0 for m in range(1, 13)},
                                               "ly": {str(m): 0 for m in range(1, 13)}})
                if series.get("cat"):
                    out["cat"] = series["cat"]
                for key in ("actual", "ly"):
                    for m in range(1, 13):
                        out[key][str(m)] = round(out[key][str(m)] + series[key][str(m)], 2)
        mt_channel[field] = merged

    # merge annual / annual_by_cat / annual_by_sku
    for field in ("annual", "annual_by_cat", "annual_by_sku"):
        merged = {}
        for ch_id in mt:
            source = chan_data[ch_id][field]
            if field == "annual":
                source = {"_total": source}   # wrap {year: val} so value is a dict
            for name, years in source.items():
                out = merged.setdefault(name, {})
                for year, value in years.items():
                    out[year] = round(out.get(year, 0) + value, 2)
        mt_channel[field] = merged["_total"] if field == "annual" else merged

    # merge history / history_by_cat / history_by_sku
    for field in ("history", "history_by_cat", "history_by_sku"):
        merged = {}
        for ch_id in mt:
            source = chan_data[ch_id][field]
            if field == "history":
                source = {"_total": source}
            for name, years in source.items():
                out = merged.setdefault(name, {})
                for year, months in years.items():
                    yy = out.setdefault(year, {})
                    for m, value in months.items():
                        yy[m] = round(yy.get(m, 0) + value, 2)
        mt_channel[field] = merged["_total"] if field == "history" else merged

    chan_data["MT"] = mt_channel

    # _channels: MT first + every 2026-active channel (total_mb>0), ordered by 2026 total desc
    #   (matches original behaviour; historical-only channels are excluded from the picker)
    active = [(chan_data[c]["total_mb"], c) for c in chan_data
              if not c.startswith("_") and c != "MT" and chan_data[c]["total_mb"] > 0]
    active.sort(key=lambda x: (-x[0], x[1]))
    chan_data["_channels"] = ([{"id": "MT", "label": mt_channel["label"]}]
                              + [{"id": c, "label": chan_data[c]["label"]} for _, c in active])

    return chan_data

def last_sales_date(last_closed):
    """Return the latest actual sales date (iso str) scoped to the dashboard's
    MT = MAKRO + LOTUS' channels, within the last closed month (last_closed).

    Date column (Raw data col 3) may be an Excel serial int or a DD/MM/YYYY
    string. The raw sheet also contains forward/other-channel rows with nonzero
    values for future months, so we must scope to MT channels AND month <= the
    last closed month derived from Dashboard Calc, otherwise we would show a
    future date that does not reflect actual MT sales.
    """
    import datetime as _dt
    _EPOCH = _dt.date(1899, 12, 30)

    def to_date(v):
        if isinstance(v, int) and v > 10000:
            return _EPOCH + _dt.timedelta(days=int(v))
        if isinstance(v, str):
            try:
                dd, mm, yy = v.strip().split("/")
                return _dt.date(int(yy), int(mm), int(dd))
            except Exception:
                return None
        return None

    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Raw data"]
    best = None
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        try:
            year = int(r[0])
        except (TypeError, ValueError):
            continue
        if year != FISCAL_YEAR:
            continue
        if norm_channel(r[5]) not in ("MAKRO", "LOTUS'"):
            continue
        d = to_date(r[3])
        if d is None:
            continue
        # Scope to the last closed month using the PARSED Date month (the raw
        # sheet's Month column buckets some rows oddly; the Date itself is
        # authoritative for "latest sales as of"). Exclude forward/future rows.
        if d.month > last_closed:
            continue
        try:
            val = float(r[10] or 0)
        except (TypeError, ValueError):
            continue
        if val == 0:
            continue
        if best is None or d > best:
            best = d
    wb.close()
    # Fall back to the last day of the last closed month if raw dates are unreadable.
    if best is None:
        best = _dt.date(FISCAL_YEAR, last_closed, 28)
    return best.isoformat()


def _load_dashboard_calc():
    """Read Dashboard Calc into per-month dicts (MB units). Used ONLY to derive
    the last-closed month and to keep the legacy MAKRO/LOTUS'/MT actual+LY series
    identical to what has been committed. LE/AOP now come from Data Plan Sale."""
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Dashboard Calc"]
    mon = {}; makro = {}; lotus = {}
    for r in ws.iter_rows(values_only=True):
        if isinstance(r[0], (int, float)) and 1 <= r[0] <= 12:
            mon[r[0]] = {"act": (r[1] or 0) + (r[3] or 0),
                         "le": (r[2] or 0) + (r[4] or 0),
                         "ly": (r[5] or 0) + (r[6] or 0),
                         "aop": (r[7] or 0) + (r[8] or 0)}
            makro[r[0]] = {"act": r[1] or 0, "le": r[2] or 0, "ly": r[5] or 0, "aop": r[7] or 0}
            lotus[r[0]] = {"act": r[3] or 0, "le": r[4] or 0, "ly": r[6] or 0, "aop": r[8] or 0}
    wb.close()
    return mon, makro, lotus


def _load_plan_sale():
    """Load Data Plan Sale -> channel -> scenario (AOP/LE) -> {month: MB}.
    All values are stored in MB units (same as Dashboard Calc / UI series).
    This is the PRIMARY LE/AOP source for all 15 channels."""
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Data Plan Sale"]
    out = {}
    for r in ws.iter_rows(values_only=True):
        if r[0] is None or not isinstance(r[4], (int, float)):
            continue
        cust = str(r[0]).strip()
        if r[4] is None:
            continue
        try:
            month = int(r[1])
        except (TypeError, ValueError):
            continue
        if not 1 <= month <= 12:
            continue
        scen = (str(r[3]) or "").strip().upper()
        if scen not in ("AOP", "LE"):
            continue
        v = float(r[4] or 0)
        out.setdefault(cust, {}).setdefault(scen, {})[month] = v
    wb.close()
    return out


def merge_kpi(d):
    """Attach planning series (actual/LE/LY/AOP) + YTD KPIs for MT and ALL 15
    channels, in MB units.

    LE and AOP are sourced PRIMARILY from the "Data Plan Sale" (tblPlanSales)
    sheet, which holds these forecast rows for every channel (15 channels x 12
    months x {AOP, LE}). Actual and LY come from each channel's own Raw-data
    monthly/history (MAKRO/LOTUS' keep the existing Dashboard Calc actual|LY so
    nothing MAKRO/Lotus-dependent regresses; those values agree with Data Plan
    Sale LE/AOP anyway). """
    mon, makro, lotus = _load_dashboard_calc()
    plan = _load_plan_sale()

    last_closed = max([m for m in mon if (mon[m]["act"] or 0) > 0], default=8)
    act_sum = sum(mon[m]["act"] for m in range(1, last_closed + 1))
    # LE mix: closed months use actual, the latest open month uses its LE forecast
    le_mix = sum(mon[m]["act"] for m in range(1, last_closed)) + mon[last_closed]["le"]
    ly_sum = sum(mon[m]["ly"] for m in range(1, last_closed + 1))
    aop_sum = sum(mon[m]["aop"] for m in range(1, last_closed + 1))

    mt = d["MT"]
    mt["actual_ytd"] = round(act_sum, 2)
    mt["le_ytd_mix"] = round(le_mix, 2)
    mt["ly_ytd"] = round(ly_sum, 2)
    mt["aop_ytd"] = round(aop_sum, 2)
    mt["s_actual"] = {str(m): round(mon[m]["act"], 2) for m in range(1, 13)}
    # MT LE/AOP from Data Plan Sale (Makro + Lotus'), matching Dashboard Calc.
    mt["s_le"] = {str(m): round(plan.get("MAKRO", {}).get("LE", {}).get(m, 0)
                              + plan.get("LOTUS'", {}).get("LE", {}).get(m, 0), 2)
                   for m in range(1, 13)}
    mt["s_ly"] = {str(m): round(mon[m]["ly"], 2) for m in range(1, 13)}
    mt["s_aop"] = {str(m): round(plan.get("MAKRO", {}).get("AOP", {}).get(m, 0)
                               + plan.get("LOTUS'", {}).get("AOP", {}).get(m, 0), 2)
                    for m in range(1, 13)}

    # MAKRO / LOTUS' keep their committed actual|LY from Dashboard Calc; their
    # LE/AOP switch to Data Plan Sale (values are identical, source is primary).
    for channel_id, series in (("MAKRO", makro), ("LOTUS'", lotus)):
        ch = d.get(channel_id)
        if ch is None:
            continue
        ch["s_actual"] = {str(m): round(series[m]["act"], 2) for m in range(1, 13)}
        ch["s_ly"] = {str(m): round(series[m]["ly"], 2) for m in range(1, 13)}
        ch["s_le"] = {str(m): round(plan.get(channel_id, {}).get("LE", {}).get(m, 0), 2)
                       for m in range(1, 13)}
        ch["s_aop"] = {str(m): round(plan.get(channel_id, {}).get("AOP", {}).get(m, 0), 2)
                        for m in range(1, 13)}

    # All OTHER channels: enable the planning chart by giving each channel its own
    # actual (Raw monthly) + LY (2025 history) and the Data Plan Sale LE/AOP.
    for channel_id in plan:
        if channel_id in ("MAKRO", "LOTUS'"):
            continue
        ch = d.get(channel_id)
        if ch is None:
            continue
        hist25 = ch.get("history", {}).get("2025", {})
        # BUGFIX 2026-08-14: ch["monthly"] and hist25 keyed by STRING months
        # ("1".."12"), but the range was looking up with int m -> always 0,
        # so other-channel s_actual/s_ly stayed all-zero, making app.js hasS()
        # gate FAIL and hide LE/AOP for non-MT channels. Look up by str(m).
        ch["s_actual"] = {str(m): round(ch.get("monthly", {}).get(str(m), 0), 2)
                           for m in range(1, 13)}
        ch["s_ly"] = {str(m): round(hist25.get(str(m), 0), 2) for m in range(1, 13)}
        ch["s_le"] = {str(m): round(plan.get(channel_id, {}).get("LE", {}).get(m, 0), 2)
                       for m in range(1, 13)}
        ch["s_aop"] = {str(m): round(plan.get(channel_id, {}).get("AOP", {}).get(m, 0), 2)
                        for m in range(1, 13)}

    d["_lastClosed"] = last_closed
    return last_closed

if __name__ == "__main__":
    data = build()
    last_closed = merge_kpi(data)
    data["_lastSalesDate"] = last_sales_date(last_closed)
    # Preserve the committed top-level key order so the git diff only shows real
    # data changes (no churn from reordering). Unknown channels keep insertion order.
    committed_order = [
        "Central food wholesales", "7Eleven", "Maxmart", "BigC", "Tops",
        "MAKRO", "LOTUS'", "MaxValue", "CJ Express", "Jiffy", "Lawson",
        "Foodland", "FamilyMart", "Golden Place", "Villa market", "MT",
    ]
    ordered = {k: data[k] for k in committed_order if k in data}
    for k in data:
        if k not in ordered:
            ordered[k] = data[k]
    data = ordered

    def _ints(o):
        """Match the committed data_channels.json convention: serialize any
        float that is a whole number as an int (0.0 -> 0, 5.0 -> 5). This keeps
        the git diff tiny and identical to the previous committed representation;
        it does not change any numeric value."""
        if isinstance(o, float):
            return int(o) if o == int(o) else o
        if isinstance(o, dict):
            return {k: _ints(v) for k, v in o.items()}
        if isinstance(o, list):
            return [_ints(v) for v in o]
        return o
    data = _ints(data)
    # Write UTF-8 (NO BOM) with 2-space indent and LF line endings.
    # IMPORTANT: keep this UTF-8. The dashboard loads this file via
    # fetch().json(), which can only parse UTF-8. A prior version wrote
    # UTF-16 LE + BOM, which broke the browser JSON parser and made the
    # dashboard show no data after deploy (fixed 2026-08-14).
    text = json.dumps(data, ensure_ascii=False, indent=2)
    with open(OUT, "w", encoding="utf-8", newline="") as fp:
        fp.write(text)
        fp.write("\n")
    mt = data["MT"]
    print("SAVED", OUT)
    print("last_closed month:", last_closed)
    print("last_sales_date:", data["_lastSalesDate"])
    print("MT annual:", mt["annual"])
    print("MT total_mb (2026 YTD):", mt["total_mb"])
    print("MT actual_ytd:", mt["actual_ytd"], "| le_ytd_mix:", mt["le_ytd_mix"],
          "| ly_ytd:", mt["ly_ytd"], "| aop_ytd:", mt["aop_ytd"])
    print("channels:", [c["id"] for c in data["_channels"]])
