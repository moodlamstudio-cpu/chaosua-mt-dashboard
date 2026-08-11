from openpyxl import load_workbook
from collections import defaultdict
import json
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
pm={}
ws=wb["Product Master"]
for i,r in enumerate(ws.iter_rows(values_only=True)):
    if i==0: continue
    pm[str(r[0])]={"d":(r[4] or ""),"g":(r[8] or ""),"c1":(r[13] or "")}
ws2=wb["Raw data"]
sku=defaultdict(lambda:[0.0,""])  # mat -> [val, cat]
for i,r in enumerate(ws2.iter_rows(values_only=True)):
    if i==0: continue
    if r[0]!=2026 or r[1] is None or r[1]>8: continue
    mm=str(r[4] or "").replace(" ","").lower()
    if mm not in ("makro","lotus's","lotus"): continue
    mat=str(r[5]); val=r[9] or 0
    info=pm.get(mat,{})
    sku[mat][0]+=val
    if sku[mat][1]=="" :
        sku[mat][1]=info.get("c1") or info.get("g") or "Other"
dp="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Web/LotusMakro-Dashboard/data.json"
d=json.load(open(dp,encoding="utf-8"))
d["top_sku"]=[{"name":k,"mb":round(v[0]/1e6,2),"cat":v[1]} for k,v in sorted(sku.items(),key=lambda x:-x[1][0])[:10]]
json.dump(d,open(dp,"w",encoding="utf-8"),ensure_ascii=False,indent=2)
print("top_sku พร้อม cat:")
for s in d["top_sku"]: print(f"  {s['mb']:6.2f}  [{s['cat']}]  {s['name']}")
wb.close()
