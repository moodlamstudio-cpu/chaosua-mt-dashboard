from openpyxl import load_workbook
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
print("Sheets:",wb.sheetnames)
# ดู Raw data header + ตัวอย่าง
if "Raw data" in wb.sheetnames:
    ws=wb["Raw data"]
    rows=ws.iter_rows(min_row=1,max_row=6,values_only=True)
    hdr=next(rows,None)
    print("Raw hdr:",hdr)
    for r in rows:
        print("  ",r)
wb.close()
