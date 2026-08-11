from openpyxl import load_workbook
from collections import defaultdict
import json
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
pm={}
ws=wb["Product Master"]
for i,r in enumerate(ws.iter_rows(values_only=True)):
    if i==0: continue
    if r[0] is None: continue
    pm[str(r[0])]={"d":(r[4] or ""),"g":(r[8] or ""),"c1":(r[13] or "")}
ws2=wb["Raw data"]
sku=defaultdict(lambda:[0.0,"",""])  # mat -> [val, desc, cat]
for i,r in enumerate(ws2.iter_rows(values_only=True)):
    if i==0: continue
    if r[0]!=2026 or r[1] is None or r[1]>8: continue
    mm=str(r[4] or "").replace(" ","").lower()
    if mm not in ("makro","lotus's","lotus"): continue
    mat=str(r[5]); val=r[9] or 0
    info=pm.get(mat,{})
    sku[mat][0]+=val
    if sku[mat][1]=="":
        sku[mat][1]=info.get("d") or mat
        sku[mat][2]=info.get("c1") or info.get("g") or "Other"
# group by cat -> sorted top 10 each
bycat=defaultdict(list)
for mat,(v,d,c) in sku.items():
    bycat[c].append({"name":d,"mb":round(v/1e6,2)})
for c in bycat: bycat[c].sort(key=lambda x:-x["mb"])
top_all=[]; all_sku=[]
# top 10 รวมทุกหมวด + per-cat top10
all_sorted=sorted(((v,mat) for mat,(v,d,c) in sku.items()),reverse=True)
for v,mat in all_sorted[:10]:
    d,c=sku[mat][1],sku[mat][2]
    top_all.append({"name":d,"mb":round(v/1e6,2),"cat":c})
percat={c:[{"name":s["name"],"mb":s["mb"]} for s in lst[:10]] for c,lst in bycat.items()}
# เอาเฉพาะหมวดหลัก (ที่มีในพาย)
d=json.load(open("data.json",encoding="utf-8"))
cats=[c["name"] for c in d["category"]["items"]]
d["top_sku"]=top_all
d["top_sku_by_cat"]={c:percat.get(c,[]) for c in cats}
json.dump(d,open("data.json","w",encoding="utf-8"),ensure_ascii=False,indent=2)
print("top_sku (รวมทั้งหมด):",len(top_all))
for c in cats:
    print(f"  [{c}] top {len(percat.get(c,[]))} SKUs")
    for s in percat.get(c,[])[:3]: print(f"      {s['name']} {s['mb']} MB")
wb.close()
