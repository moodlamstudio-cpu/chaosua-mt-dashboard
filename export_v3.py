from openpyxl import load_workbook
import json, datetime
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,data_only=True)
ws=wb["Dashboard Calc"]
rows=list(ws.iter_rows(min_row=4,max_row=17,values_only=True))
hdr=[(str(c).strip() if c is not None else "") for c in rows[0]]
def F(r,k):
    if k not in hdr: return 0
    v=r[hdr.index(k)]
    return round(float(v),1) if v not in (None,"") and not(isinstance(v,str) and "#" in v) else 0
months=[]; ytd=None
for r in rows[1:]:
    mn=str(r[0]).strip()
    if mn.upper()=="YTD": ytd=r; continue
    mi=int(float(mn))
    months.append({"m":mi,"label":{1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}[mi],
      "makro":{"actual":F(r,"Makro Actual"),"le":F(r,"Makro LE"),"ly":F(r,"Makro LY"),"aop":F(r,"Makro AOP")},
      "lotus":{"actual":F(r,"Lotus Actual"),"le":F(r,"Lotus LE"),"ly":F(r,"Lotus LY"),"aop":F(r,"Lotus AOP")},
      "total":{"actual":F(r,"Actual Total"),"le":F(r,"LE Total"),"ly":F(r,"LY Total"),"aop":F(r,"AOP Total")}})
closed=max([m["m"] for m in months if m["total"]["actual"]>0], default=0)
out={"title":"MT Sales Dashboard","subtitle":"Lotus's + Makro (MT) · Monthly LE vs LY vs AOP",
 "year":2026,"current_month":8,"unit":"ล้านบาท","closed_month":closed,"months":months,
 "ytd":{"makro_ly":round(F(ytd,"Makro LY"),1),"makro_aop":round(F(ytd,"Makro AOP"),1),"makro_actual":round(F(ytd,"Makro Actual"),1),"makro_le":round(F(ytd,"Makro LE"),1),
   "lotus_ly":round(F(ytd,"Lotus LY"),1),"lotus_aop":round(F(ytd,"Lotus AOP"),1),"lotus_actual":round(F(ytd,"Lotus Actual"),1),"lotus_le":round(F(ytd,"Lotus LE"),1),
   "total_ly":round(F(ytd,"LY Total"),1),"total_aop":round(F(ytd,"AOP Total"),1),"total_actual":round(F(ytd,"Actual Total"),1),"total_le":round(F(ytd,"LE Total"),1)},
 "updated":datetime.date.today().strftime("%Y-%m-%d")}
with open("C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Web/LotusMakro-Dashboard/data.json","w",encoding="utf-8") as fp:
    json.dump(out,fp,ensure_ascii=False,indent=2)
print("OK closed=",closed,"actual_M1-7+LE", sum(m["total"]["actual"] for m in months[:7])+sum(m["total"]["le"] for m in months[7:]))
wb.close()
