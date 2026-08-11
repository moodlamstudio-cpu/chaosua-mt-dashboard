from openpyxl import load_workbook
import json
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
ws=wb["Dashboard Calc"]
rows=list(ws.iter_rows(values_only=True))
mon={}
makro={}
lotus={}
for r in rows:
    if isinstance(r[0],(int,float)) and 1<=r[0]<=12:
        mon[r[0]]={"act":(r[1] or 0)+(r[3] or 0),"le":(r[2] or 0)+(r[4] or 0),"ly":(r[5] or 0)+(r[6] or 0),"aop":(r[7] or 0)+(r[8] or 0)}
        makro[r[0]]={"act":r[1] or 0,"le":r[2] or 0,"ly":r[5] or 0,"aop":r[7] or 0}
        lotus[r[0]]={"act":r[3] or 0,"le":r[4] or 0,"ly":r[6] or 0,"aop":r[8] or 0}
# เลขเป็นล้านบาทตรงๆ (ไม่ ÷)
act=sum(mon[m]["act"] for m in range(1,9))                     # 104.35
le_mix=sum(mon[m]["act"] for m in range(1,8))+mon[8]["le"]     # Actual M1-7 + LE M8
ly=sum(mon[m]["ly"] for m in range(1,9))
aop=sum(mon[m]["aop"] for m in range(1,9))
print(f"MT KPI: Actual={act:.2f} LE(mix)={le_mix:.2f} LY={ly:.2f} AOP={aop:.2f} vsAOP={le_mix/aop*100:.1f}%")
# monthly series for chart (เป็น MB อยู่แล้ว)
for m in [1,4,8]:
    print(f"  M{m}: act={mon[m]['act']:.2f} le={mon[m]['le']:.2f} ly={mon[m]['ly']:.2f} aop={mon[m]['aop']:.2f}")
# เขียน data_channels.json
d=json.load(open("data_channels.json",encoding="utf-8"))
mt=d["MT"]
mt["actual_ytd"]=round(act,2)
mt["le_ytd_mix"]=round(le_mix,2)
mt["ly_ytd"]=round(ly,2)
mt["aop_ytd"]=round(aop,2)
mt["s_actual"]={str(m):round(mon[m]["act"],2) for m in range(1,13)}
mt["s_le"]={str(m):round(mon[m]["le"],2) for m in range(1,13)}
mt["s_ly"]={str(m):round(mon[m]["ly"],2) for m in range(1,13)}
mt["s_aop"]={str(m):round(mon[m]["aop"],2) for m in range(1,13)}
# MAKRO และ LOTUS' มี LE/LY/AOP แยกช่องทางจาก Dashboard Calc
for channel_id, series in (("MAKRO",makro),("LOTUS'",lotus)):
    ch=d[channel_id]
    ch["s_actual"]={str(m):round(series[m]["act"],2) for m in range(1,13)}
    ch["s_le"]={str(m):round(series[m]["le"],2) for m in range(1,13)}
    ch["s_ly"]={str(m):round(series[m]["ly"],2) for m in range(1,13)}
    ch["s_aop"]={str(m):round(series[m]["aop"],2) for m in range(1,13)}
json.dump(d,open("data_channels.json","w",encoding="utf-8"),ensure_ascii=False,indent=2)
print("บันทึกแล้ว: actual_ytd=",mt["actual_ytd"],"le_mix=",mt["le_ytd_mix"],"ly=",mt["ly_ytd"],"aop=",mt["aop_ytd"])
wb.close()
