from openpyxl import load_workbook
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
wb=load_workbook(f,read_only=True,data_only=True)
# 1) Dashboard Calc - ดูค่าจริง
ws=wb["Dashboard Calc"]
print("=== Dashboard Calc (rows 4-17) col: Month,MakroAct,LotusAct ===")
for i,r in enumerate(ws.iter_rows(values_only=True)):
    if i<3 or i>16: continue
    if r[0] is None: continue
    print(f"  {r[0]}: MakroAct={r[1]} LotusAct={r[3]}")
# 2) Raw data เดือน1-8 กรอง Makro+Lotus เทียบ
raw={}
for i,r in enumerate(wb["Raw data"].iter_rows(values_only=True)):
    if i==0: continue
    if r[0]!=2026: continue
    m=r[1]
    if m is None or m>8: continue
    c=str(r[4] or "").replace(" ","").lower()
    if c not in ("makro","lotus","lotus's"):
        # print นอกขอบเขต นับรวมเผื่อ
        pass
    else:
        k=(m,c)
        raw[k]=raw.get(k,0)+(r[9] or 0)
# sum
tot=sum(v for (m,c),v in raw.items())
print(f"\nRaw data 2026 M1-8 Makro+Lotus = {tot/1e6:.2f} MB")
# แยก
m=sum(v for (mm,c),v in raw.items() if c=="makro")
l=sum(v for (mm,c),v in raw.items() if c in ("lotus","lotus's"))
print(f"  Makro={m/1e6:.2f} Lotus={l/1e6:.2f}")
wb.close()
