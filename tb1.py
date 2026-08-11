import openpyxl
print("openpyxl", openpyxl.__version__)
from openpyxl import load_workbook
f="C:/Users/teera/OneDrive - TIA NGEE HIANG (CHAOSUA) CO.,LTD/Desktop/Chaosua_Ice/Data/Sales report/Sale Lotus Makro_Dashboard_Pivot.xlsx"
# data_only=True อ่านค่าที่คำนวณ (cache)
wb=load_workbook(f,read_only=True,data_only=True)
ws=wb["Dashboard Calc"]
# อ่านทุกแถวด้วย values_only ดิบ ดูว่า r[1],r[5],r[7] คืนค่าจริงไหม (M1)
cnt=0
for r in ws.iter_rows(values_only=True):
    if r[0] in (1,2):  # M1 M2
        print("M",r[0],"| act_total:",r[9],"| makro_act:",r[1],"| makro_ly:",r[5],"| lotus_ly:",r[6],"| makro_aop:",r[7],"| lotus_aop:",r[8])
        cnt+=1
    if cnt>=2: break
# ตรวจว่าค่าเป็น None หรือ 0 (read cache หาย?)
for r in ws.iter_rows(values_only=True):
    if r[0]==1:
        print("M1 raw cells:", [ (type(x).__name__, x) for x in r[:12] ])
        break
wb.close()
